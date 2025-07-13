import itertools
import re
import string
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional, Tuple

import numpy as np
import openpyxl.cell.rich_text
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet

TYPE_WORKSHEET = Worksheet


class ExcelEditor:

    def __init__(self, path_excel: Path):
        self.wb = load_workbook(path_excel, rich_text=True)

    def get_pages_name(self) -> List[str]:
        return [ws.title for ws in self.wb.worksheets]

    def get_page(self, page_name: str) -> TYPE_WORKSHEET:
        lst = [ws for ws in self.wb.worksheets if ws.title == page_name]
        if len(lst) != 1:
            raise ValueError(
                f"The name page '{page_name}' has {len(lst)} corresponding worksheets."
            )
        return lst[0]

    def read_cell(self, page_name: str, idx_row: int, idx_column: int) -> Optional[str]:
        ws = self.get_page(page_name)
        return ws.cell(idx_row, idx_column).value

    @staticmethod
    def convert_column_to_number(column_name: str) -> int:

        numbers = "".join(str(i) for i in range(1, 10))
        to_base_char = dict(
            zip(string.ascii_uppercase, numbers + string.ascii_uppercase[:17])
        )

        s = "".join(to_base_char[c] for c in column_name)
        return int(s, 26)

    @staticmethod
    def get_page_dimensions(worksheet: TYPE_WORKSHEET) -> Tuple[int, int]:
        dims = worksheet.calculate_dimension()
        searches = re.search(pattern=r"[A-Z]+[0-9]+:([A-Z]+[0-9]+)", string=dims)
        nrows, ncols = coordinate_to_tuple(searches.group(1))
        return nrows, ncols

    @staticmethod
    def from_cell_to_obj(cell_value: Any) -> Optional[Any]:
        if cell_value is None:
            obj = None
        elif (
            isinstance(cell_value, str)
            or isinstance(cell_value, datetime)
            or isinstance(cell_value, bool)
            or isinstance(cell_value, float)
            or isinstance(cell_value, int)
        ):
            obj = cell_value
        elif isinstance(cell_value, CellRichText):
            obj = "".join(e if isinstance(e, str) else e.text for e in cell_value)
        else:
            raise RuntimeError(f"{type(cell_value)} not handled.")

        return obj

    def read_page(self, page_name: str) -> np.ndarray:
        page = self.get_page(page_name)
        nrows, ncols = ExcelEditor.get_page_dimensions(page)
        data = []
        for row in range(1, nrows + 1):
            lst = []
            for col in range(1, ncols + 1):
                cell_value = page.cell(row, col).value
                lst.append(ExcelEditor.from_cell_to_obj(cell_value))
            data.append(lst)

        return np.array(data)

    def save(self, path_excel) -> None:
        self.wb.save(path_excel)


if __name__ == "__main__":
    from vars import PATH_DOCS_PLANNING_MAY

    ee = ExcelEditor(PATH_DOCS_PLANNING_MAY)
    res = ee.read_page("Anglos")
    print(res[6:, 0])
